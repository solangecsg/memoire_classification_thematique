"""
composition_periode.py : composition thématique du corpus par tranche de vingt ans

CE QUE FAIT CE SCRIPT

Agrège les affectations d'un modèle thématique par période de publication, et
rend le tableau que reprend le mémoire. Aucun entraînement n'est refait :
l'affectation des documents porte l'identifiant du fascicule, et il suffit de
rattacher chaque fascicule à sa date.

D'OÙ VIENNENT LES DATES

Chaque fascicule du corpus est accompagné de son manifeste, dont l'élément
dc:date porte la date de publication. La date est donc lue dans le corpus
lui-même plutôt que dans une source extérieure, ce qui rend le calcul
reproductible à partir des seuls fichiers versés.

Trois formes se rencontrent dans cet élément et sont toutes trois reconnues :
la date simple, la date écrite avec des barres obliques, et l'intervalle de deux
dates, dont la première est retenue.

Un contrôle facultatif confronte ces dates au tableur des URLs Gallica, qui
recense les cent fascicules avec leur titre et leur date. Les deux sources
s'accordent sur les cent, les seules divergences relevées tenant à l'écriture.

POURQUOI CE SCRIPT EXISTE

Une première version du relevé ne portait que sur soixante-sept fascicules, le
tiers restant étant réputé sans date. Le contrôle a montré que les cent dates
étaient disponibles, dans le manifeste comme dans le tableur. Le présent script
refait le calcul sur l'ensemble du corpus.

ENTRÉES

  {corpus}/{fascicule}_reocr/manifest.xml                  dates de publication
  resultats/lda_corpus_bnf_article_k10_g1_*/span_topic.json  affectations
  resultats/lda_corpus_bnf_article_k10_g1_*/topics.json      mots de tête
  ../../100_urls_gallica_recap.xlsx                        contrôle facultatif

SORTIES

  Un tableau imprimé sur la sortie standard, en LaTeX avec --latex.

PAQUETS EMPLOYÉS

  openpyxl                          lecture du tableur de contrôle, facultative
  json, re, glob, collections, argparse, pathlib   bibliothèque standard

Le manifeste est lu par expression régulière plutôt que par un analyseur XML.
Un seul élément est cherché, dont la forme est stable, et cette lecture évite
d'analyser cent fichiers de cent kilooctets pour en extraire une ligne.

USAGE

    python3 composition_periode.py
    python3 composition_periode.py --latex
    python3 composition_periode.py --controle
"""

import argparse
import collections
import glob
import json
import re
from pathlib import Path

ICI = Path(__file__).resolve().parent
SORTIE = ICI / "composition"

# Tranches de vingt ans. Le corpus s'étend de 1819 à 1953, et ce pas donne des
# effectifs comparables d'une tranche à l'autre sans lisser les mouvements que
# le mémoire commente.
PAS = 20
DEBUT = 1800

# Intitulés donnés aux dix thèmes du modèle repris dans le mémoire. Ils sont
# lus par un humain sur les mots de tête, opération que le chapitre décrit comme
# le point faible de la méthode : ils servent à la lecture du tableau et ne sont
# produits par aucun calcul.
NOMS = {0: "locale", 1: "bruit", 2: "sport", 3: "bourse", 4: "annonces",
        5: "guerre", 6: "politique", 7: "spect.", 8: "divers", 9: "chron."}
ORDRE = ["locale", "bruit", "sport", "bourse", "annonces", "guerre",
         "politique", "spect.", "chron.", "divers"]


def fascicules() -> Path:
    """Dossier qui contient les fascicules ré-océrisés, un par sous-dossier.

    Deux emplacements sont possibles selon le contexte. Dans le dossier de
    travail, les fascicules se trouvent sous resultats_mistral/. Dans le dépôt,
    ils sont versés sous re-ocr/corpus/reocr_mistral/. Les deux portent la même
    structure interne, un dossier {identifiant}_reocr contenant son manifeste et
    son sous-dossier ocr/, et le premier qui existe est retenu.
    """
    for base in [ICI, *ICI.parents]:
        for rel in (Path("resultats_mistral"),
                    Path("re-ocr") / "corpus" / "reocr_mistral"):
            if (base / rel).is_dir():
                return base / rel
    raise SystemExit(
        "fascicules ré-océrisés introuvables. Ils sont attendus sous "
        "resultats_mistral/ dans le dossier de travail, ou sous "
        "re-ocr/corpus/reocr_mistral/ dans le dépôt.")


def dates():
    """Rend la date de publication de chaque fascicule, lue dans son manifeste.

    L'élément dc:date accepte trois écritures dans ce corpus. La date simple,
    2018-11-07. La date à barres obliques, 1899/12/31. Et l'intervalle,
    1953-03-01/1953-03-02, pour un fascicule couvrant plusieurs jours, dont la
    première date est retenue puisque c'est celle de parution.
    """
    out = {}
    for d in sorted(fascicules().glob("*_reocr")):
        fid = d.name.split("_")[0]
        f = d / "manifest.xml"
        if not f.is_file():
            continue
        t = f.read_text(encoding="utf-8", errors="ignore")
        m = re.search(r"<dc:date[^>]*>([^<]+)</dc:date>", t)
        if not m:
            continue
        brut = m.group(1).strip().split("/") if "/" not in m.group(1)[:8] \
            else [m.group(1).strip()]
        premiere = m.group(1).strip().split("/")[0] if m.group(1).count("/") > 1 \
            else m.group(1).strip().split("/")[0]
        an = re.match(r"(1[89]\d{2}|20\d{2})", premiere.replace("/", "-"))
        if an:
            out[fid] = int(an.group(1))
    return out


def controle(par_fascicule):
    """Confronte les dates du manifeste au tableur des URLs Gallica.

    Le tableur donne l'ARK de chaque fascicule, dont les sept chiffres qui
    suivent bpt6k forment l'identifiant employé dans le corpus, la dernière
    position de l'ARK étant un caractère de contrôle.
    """
    try:
        import openpyxl
    except ImportError:
        return "openpyxl absent, contrôle non conduit"
    cands = [c / "100_urls_gallica_recap.xlsx" for c in fascicules().parents]
    f = next((c for c in cands if c.is_file()), None)
    if f is None:
        return "tableur absent, contrôle non conduit"
    w = openpyxl.load_workbook(f)
    tab = {}
    for _, _titre, date, url in w[w.sheetnames[0]].iter_rows(min_row=2, values_only=True):
        m = re.search(r"bpt6k(\d{7})", str(url))
        if m:
            tab[m.group(1)] = int(str(date).split("/")[-1])
    communs = set(tab) & set(par_fascicule)
    acc = sum(1 for k in communs if tab[k] == par_fascicule[k])
    return (f"{len(tab)} fascicules au tableur, {len(communs)} communs, "
            f"{acc} années identiques")


def affectations():
    """Rend le fascicule et le thème de chaque document, pour le run du mémoire."""
    dossiers = sorted(glob.glob(str(ICI / "resultats" /
                                    "lda_corpus_bnf_article_k10_g1_2*")))
    if not dossiers:
        raise SystemExit("exécution introuvable : resultats/lda_corpus_bnf_article_k10_g1_*")
    run = dossiers[0]
    f = Path(run, "span_topic.json")
    if not f.is_file():
        raise SystemExit(
            f"affectation des documents absente : {f.name}\n"
            "Elle n'est pas versée dans le dépôt, étant volumineuse et "
            "régénérable. La reconstituer par :\n"
            "    python3 lda_mallet_corpus.py --source bnf --granularite article "
            "--k 10 --graine 1")
    d = json.loads(f.read_text(encoding="utf-8"))
    if isinstance(d, dict):
        d = d.get("spans", list(d.values()))
    return [(x["fascicule"], int(x["topic_id"])) for x in d
            if isinstance(x, dict) and x.get("fascicule") is not None]


# Couleurs des dix bandes. Le thème de bruit reçoit un gris, sa variation
# mesurant l'état des fichiers plutôt qu'un contenu, et les trois thèmes que le
# mémoire commente sont placés au bas de la pile, où une bande se lit contre une
# ligne de base droite plutôt que contre le sommet ondulant de celles d'en
# dessous.
COULEURS = {
    "sport":     "#c0724a",
    "spect.":    "#cf8fa0",
    "bourse":    "#4f7a5b",
    "bruit":     "#b0b0b0",
    "locale":    "#6a8caf",
    "annonces":  "#d8b25e",
    "guerre":    "#8a5a72",
    "politique": "#3f5f8a",
    "chron.":    "#7e9aa6",
    "divers":    "#d9d2c5",
}
PILE = ["sport", "spect.", "bourse", "bruit", "locale", "annonces",
        "guerre", "politique", "chron.", "divers"]

# Une bande porte son intitulé quand elle est assez épaisse pour l'accueillir,
# ce qui évite une légende que l'œil devrait faire l'aller-retour pour lire.
EPAISSEUR_INTITULE = 9.0


def tableau():
    """Croise les périodes et les thèmes, en pourcentage des articles de la période."""
    an = dates()
    docs = affectations()
    sans = sorted({f for f, _ in docs} - set(an))
    par_periode = collections.defaultdict(collections.Counter)
    for f, t in docs:
        if f in an:
            par_periode[(an[f] - DEBUT) // PAS][t] += 1
    lignes = []
    for tranche in sorted(par_periode):
        c = par_periode[tranche]
        n = sum(c.values())
        deb = DEBUT + tranche * PAS
        lignes.append((f"{deb}-{deb + PAS - 1}", n,
                       {NOMS[t]: round(100 * v / n) for t, v in c.items()}))
    return lignes, len(an), sans, sum(n for _, n, _ in lignes), len(docs)


def parts_exactes():
    """Rend les parts non arrondies, pour que la pile ferme exactement à cent.

    Le tableau du mémoire arrondit chaque case à l'unité, de sorte qu'une ligne
    peut totaliser 102. Une aire empilée qui reprendrait ces valeurs laisserait
    un bord supérieur irrégulier sans signification.
    """
    an = dates()
    par_periode = collections.defaultdict(collections.Counter)
    for f, t in affectations():
        if f in an:
            par_periode[(an[f] - DEBUT) // PAS][t] += 1
    out = []
    for tranche in sorted(par_periode):
        c = par_periode[tranche]
        n = sum(c.values())
        deb = DEBUT + tranche * PAS
        out.append((deb + PAS // 2, f"{deb}-{deb + PAS - 1}", n,
                    {NOMS[t]: 100 * v / n for t, v in c.items()}))
    return out


def figure():
    """Trace la composition par période en aires empilées.

    Deux panneaux partagent l'axe du temps. Celui du haut donne l'effectif de
    chaque tranche, que la pile normalisée à cent effacerait sans cela : la
    première tranche repose sur 37 articles et la sixième sur plus de mille, et
    le mémoire refuse de conclure sur la première pour cette raison. Celui du bas
    porte les dix bandes.
    """
    import matplotlib.pyplot as plt
    from matplotlib.patches import Patch

    police()
    d = parts_exactes()
    x = [a for a, _, _, _ in d]
    n = [e for _, _, e, _ in d]
    series = {k: [p.get(k, 0.0) for _, _, _, p in d] for k in PILE}

    fig, (haut, bas) = plt.subplots(
        2, 1, figsize=(6.6, 4.6), sharex=True,
        gridspec_kw={"height_ratios": [1, 5.2], "hspace": 0.10})

    haut.plot(x, n, "o-", color="#8c8c8c", lw=1.1, ms=3.2)
    for i, (xi, ni) in enumerate(zip(x, n)):
        place = "center" if 0 < i < len(x) - 1 else ("left" if i == 0 else "right")
        haut.text(xi, ni + max(n) * 0.10, f"{ni:,}".replace(",", "\u202f"),
                  ha=place, va="bottom", fontsize=6.4, color="#4d4d4d")
    haut.set_ylim(0, max(n) * 1.55)
    haut.set_ylabel("articles", fontsize=7.4, labelpad=6)
    haut.set_yticks([])
    for c in ("top", "right", "left"):
        haut.spines[c].set_visible(False)
    haut.tick_params(axis="x", length=0)
    haut.set_xlim(min(x), max(x))

    bas.stackplot(x, [series[k] for k in PILE],
                  colors=[COULEURS[k] for k in PILE], linewidth=0.35,
                  edgecolor="white")

    # Intitulés posés au centre de la bande, à l'abscisse où elle est la plus
    # épaisse, et seulement là où la place suffit.
    cumul = [0.0] * len(x)
    for k in PILE:
        v = series[k]
        interieur = max(range(1, len(v) - 1), key=lambda j: v[j])
        bord = max(range(len(v)), key=lambda j: v[j])
        i = bord if v[bord] > v[interieur] + 1.5 else interieur
        if v[i] >= EPAISSEUR_INTITULE:
            # Sur une tranche d'extrémité, l'intitulé est ramené vers l'intérieur
            # et aligné sur son bord, faute de quoi il déborderait du cadre.
            place = "center" if 0 < i < len(x) - 1 else ("left" if i == 0 else "right")
            decal = {"center": 0, "left": 1.5, "right": -1.5}[place]
            clair = k in ("bruit", "divers", "annonces")
            bas.text(x[i] + decal, cumul[i] + v[i] / 2, k, ha=place, va="center",
                     fontsize=7.2, color="#333333" if clair else "white")
        cumul = [c + w for c, w in zip(cumul, v)]

    bas.set_xlim(min(x), max(x))
    bas.set_ylim(0, 100)
    bas.set_xticks(x)
    bas.set_xticklabels([p.replace("-", "--") for _, p, _, _ in d],
                        fontsize=7, rotation=30, ha="right")
    bas.set_yticks([0, 25, 50, 75, 100])
    bas.set_yticklabels(["0", "25", "50", "75", "100 %"], fontsize=7.4)
    bas.set_ylabel("part des articles de la période", fontsize=8)
    for c in ("top", "right"):
        bas.spines[c].set_visible(False)

    absents = [k for k in PILE if max(series[k]) < EPAISSEUR_INTITULE]
    if absents:
        bas.legend(handles=[Patch(facecolor=COULEURS[k], label=k) for k in absents],
                   fontsize=6.8, frameon=False, ncol=len(absents),
                   loc="lower center", bbox_to_anchor=(0.5, -0.42))

    SORTIE.mkdir(exist_ok=True)
    fig.savefig(SORTIE / "composition_periode.pdf", bbox_inches="tight")
    plt.close(fig)
    return d, absents


def figure_petits():
    """Trace un petit graphique par thème, tous à la même échelle.

    L'aire empilée montre une composition et se prête mal au suivi d'une série
    isolée, la bande du bas étant la seule à se lire contre une ligne de base
    droite. Dix petits graphiques répondent à l'autre question, celle de
    l'apparition et de la disparition d'une matière.

    Trois partis pris. L'échelle des ordonnées est commune, de sorte que les
    amplitudes se comparent d'un panneau à l'autre. Chaque panneau porte en fond
    les neuf autres séries en gris clair, ce qui situe le thème dans le champ.
    Et les panneaux sont rangés par période de maximum, si bien que la grille se
    lit du début du siècle à la fin.
    """
    import matplotlib.pyplot as plt
    import numpy as np

    police()
    d = parts_exactes()
    x = [a for a, _, _, _ in d]
    series = {k: [p.get(k, 0.0) for _, _, _, p in d] for k in PILE}
    ordre = sorted(PILE, key=lambda k: (series[k].index(max(series[k])), -max(series[k])))
    haut = max(max(v) for v in series.values())

    # Un bandeau d'effectifs coiffe la grille, comme sur l'aire empilée : la part
    # se lit de la même façon sur trente-sept articles et sur mille trois cents,
    # et le texte refuse de conclure sur la première tranche pour cette raison.
    fig = plt.figure(figsize=(6.8, 3.5))
    grille = fig.add_gridspec(3, 5, height_ratios=[0.55, 2.1, 2.1],
                              hspace=0.55, wspace=0.38)
    bandeau = fig.add_subplot(grille[0, :])
    n = [e for _, _, e, _ in d]
    bandeau.plot(x, n, "o-", color="#8c8c8c", lw=1.0, ms=2.8)
    for i, (xi, ni) in enumerate(zip(x, n)):
        place = "center" if 0 < i < len(x) - 1 else ("left" if i == 0 else "right")
        bandeau.text(xi, ni + max(n) * 0.12, f"{ni:,}".replace(",", "\u202f"),
                     ha=place, va="bottom", fontsize=5.8, color="#4d4d4d")
    bandeau.set_ylim(0, max(n) * 1.7)
    bandeau.set_xlim(min(x), max(x))
    bandeau.set_ylabel("articles", fontsize=6.6, labelpad=4)
    bandeau.set_yticks([])
    bandeau.set_xticks([])
    for c in ("top", "right", "left"):
        bandeau.spines[c].set_visible(False)

    axes = np.array([[fig.add_subplot(grille[r + 1, c]) for c in range(5)]
                     for r in range(2)])
    for ax, k in zip(axes.flat, ordre):
        for autre in PILE:
            if autre != k:
                ax.plot(x, series[autre], color="#e8e8e8", lw=0.5, zorder=1)
        ax.fill_between(x, series[k], color=COULEURS[k], alpha=0.85, lw=0, zorder=2)
        ax.plot(x, series[k], color=COULEURS[k], lw=1.0, zorder=3)
        i = series[k].index(max(series[k]))
        ax.set_title(k, fontsize=7.6, pad=3)
        place = "center" if 0 < i < len(x) - 1 else ("left" if i == 0 else "right")
        ax.annotate(f"{max(series[k]):.0f}", xy=(x[i], max(series[k])),
                    xytext=(0, 2), textcoords="offset points", fontsize=6,
                    color="#4d4d4d", ha=place, va="bottom")
        ax.set_ylim(0, haut * 1.16)
        ax.set_xlim(min(x), max(x))
        ax.grid(True, axis="y", lw=0.3, color="#ededed")
        ax.set_axisbelow(True)
        for c in ("top", "right"):
            ax.spines[c].set_visible(False)
        ax.tick_params(labelsize=6.2, length=2)
        ax.set_xticks([])
        ax.set_yticks([])

    for ax in axes[1]:
        ax.set_xticks([x[0], x[3], x[-1]])
        ax.set_xticklabels([str(v) for v in (x[0], x[3], x[-1])], fontsize=5.8)
    for ax in axes[:, 0]:
        ax.set_yticks([0, 20, 40])
        ax.tick_params(labelsize=6.2)
    axes[0, 0].set_ylabel("part des articles", fontsize=7)
    axes[1, 0].set_ylabel("part des articles", fontsize=7)

    SORTIE.mkdir(exist_ok=True)
    fig.savefig(SORTIE / "composition_petits.pdf", bbox_inches="tight")
    plt.close(fig)
    return ordre, [(k, max(series[k]), d[series[k].index(max(series[k]))][1]) for k in ordre]


def police():
    """Charge la fonte du mémoire, par la fonction du script des projections."""
    import importlib.util
    spec = importlib.util.spec_from_file_location(
        "projections_clusters", ICI / "projections_clusters.py")
    m = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(m)
    m.police()


def main():
    p = argparse.ArgumentParser(description=__doc__.split("\n")[1])
    p.add_argument("--latex", action="store_true", help="écrire le corps du tableau LaTeX")
    p.add_argument("--controle", action="store_true", help="confronter au tableur Gallica")
    p.add_argument("--figure", action="store_true",
                   help="tracer la composition en aires empilées")
    p.add_argument("--petits", action="store_true",
                   help="tracer un petit graphique par thème, à échelle commune")
    args = p.parse_args()

    if args.petits:
        ordre, pics = figure_petits()
        print(f"composition/composition_petits.pdf : {len(ordre)} panneaux, "
              "rangés par période de maximum")
        for k, v, periode in pics:
            print(f"  {k:<10} maximum {v:>5.1f}% en {periode}")
        return

    if args.figure:
        d, absents = figure()
        print(f"composition/composition_periode.pdf : {len(d)} périodes, "
              f"{sum(e for _, _, e, _ in d)} articles")
        for _, periode, e, parts in d:
            tete = sorted(parts.items(), key=lambda kv: -kv[1])[:3]
            print(f"  {periode}  n={e:<5} " +
                  "  ".join(f"{k} {v:.0f}%" for k, v in tete))
        if absents:
            print("  bandes trop minces pour porter leur intitulé, mises en "
                  "légende : " + ", ".join(absents))
        return

    lignes, n_dates, sans, couverts, total = tableau()
    if args.controle:
        print(controle(dates()))
        return
    if args.latex:
        for periode, n, parts in lignes:
            cases = " & ".join(f"{parts.get(k, 0)}" for k in ORDRE)
            print(f"{periode} & {n:<5} & {cases} \\\\")
        return
    print(f"fascicules datés : {n_dates}")
    print(f"articles couverts : {couverts} sur {total}")
    if sans:
        print(f"fascicules sans date : {', '.join(sans)}")
    entete = f"{'Période':<10}{'n':>6}  " + "".join(f"{k:>10}" for k in ORDRE)
    print("\n" + entete)
    for periode, n, parts in lignes:
        print(f"{periode:<10}{n:>6}  " + "".join(f"{parts.get(k, 0):>10}" for k in ORDRE))


if __name__ == "__main__":
    main()
